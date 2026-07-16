# Local module imports
# Standard library imports
import os
import sys
import tempfile
from typing import Any, Tuple

import pandas as pd  # type: ignore[import-untyped]

try:
    import win32com.client as win32com_client  # type: ignore[import-not-found]
except ModuleNotFoundError:
    win32com_client = None  # type: ignore[assignment]

# Document processing libraries
from docx import Document as WordDoc
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from pptx import Presentation

from Commons.Exceptions import DataProcessingError
from Commons.SingletonMixin import SingletonMixin
from Config.Config import Config
from Gui.PrettyWriter import PrettyWriter
from Helpers.FileUtils import FileUtils


class OfficeDocConverter(SingletonMixin):
    """
    Singleton class that converts legacy Office files (.doc, .ppt, .xls)
    into modern formats (.docx, .pptx, .xlsx), loads them into memory,
    and provides utility functions (e.g., worksheet → DataFrame).
    """

    @staticmethod
    def is_windows_supported() -> bool:
        """Return True only on Windows hosts with win32com available."""
        return sys.platform == "win32" and win32com_client is not None

    def _require_windows_support(self) -> None:
        if not self.is_windows_supported():
            raise DataProcessingError(
                "Legacy Office conversion requires Windows COM support (win32com)."
            )

    def __init__(
        self, *, cfg: "Config | None" = None, pretty: "PrettyWriter | None" = None
    ) -> None:
        # Prevent re-initialization on subsequent instantiations
        if self._initialized:
            return
        self._initialized = True

        self.pretty: PrettyWriter = pretty or PrettyWriter()
        self.fileUtils: FileUtils = FileUtils()
        self.converted_obj: Any = None
        self.file_path: str = ""

    def doc_to_docx(self) -> Any:
        try:
            self._require_windows_support()
            assert win32com_client is not None
            self.pretty.write("I", "+", f"Converting .doc → .docx: {self.file_path}")
            word = win32com_client.Dispatch("Word.Application")
            word.Visible = False

            tmp_path = self.fileUtils.randomTempFilename(".docx")
            doc_win = word.Documents.Open(self.file_path)
            doc_win.SaveAs(tmp_path, FileFormat=16)  # wdFormatXMLDocument
            doc_win.Close()
            word.Quit()

            self.converted_obj = WordDoc(tmp_path)
            os.remove(tmp_path)
            self.pretty.write(
                "O", "+", f"Converted and loaded DOCX from {self.file_path}"
            )
            return self.converted_obj

        except Exception as e:
            self.pretty.write("E", "+", f"DOC → DOCX conversion failed: {e}")
            raise DataProcessingError(f"DOC → DOCX conversion failed: {e}") from e

    def ppt_to_pptx(self, keep_app_open: bool = False) -> Any:
        try:
            self._require_windows_support()
            assert win32com_client is not None
            self.pretty.write("I", "+", f"Converting .ppt → .pptx: {self.file_path}")
            ppt_app = win32com_client.Dispatch("PowerPoint.Application")
            ppt_app.Visible = True
            ppt_app.WindowState = 2  # Minimize

            with tempfile.NamedTemporaryFile(suffix=".pptx", delete=False) as tmp:
                tmp_path = tmp.name

            pres = ppt_app.Presentations.Open(self.file_path, WithWindow=False)
            pres.SaveAs(tmp_path, 24)  # ppSaveAsOpenXMLPresentation
            pres.Close()

            if not keep_app_open:
                ppt_app.Quit()
            else:
                self.pretty.write("A", "+", "PowerPoint application kept open")

            self.converted_obj = Presentation(tmp_path)
            os.remove(tmp_path)
            self.pretty.write(
                "O", "+", f"Converted and loaded PPTX from {self.file_path}"
            )
            return self.converted_obj

        except Exception as e:
            self.pretty.write("E", "+", f"PPT → PPTX conversion failed: {e}")
            raise DataProcessingError(f"PPT → PPTX conversion failed: {e}") from e

    def xls_to_xlsx(self) -> Any:
        try:
            self._require_windows_support()
            assert win32com_client is not None
            self.pretty.write("I", "+", f"Converting .xls → .xlsx: {self.file_path}")
            excel = win32com_client.Dispatch("Excel.Application")
            excel.Visible = False

            tmp_path = self.fileUtils.randomTempFilename(".xlsx")
            wb = excel.Workbooks.Open(self.file_path)
            wb.SaveAs(tmp_path, FileFormat=51)  # xlOpenXMLWorkbook
            wb.Close()
            excel.Quit()

            self.converted_obj = load_workbook(tmp_path)
            os.remove(tmp_path)
            self.pretty.write(
                "O", "+", f"Converted and loaded XLSX from {self.file_path}"
            )
            return self.converted_obj

        except Exception as e:
            self.pretty.write("E", "+", f"XLS → XLSX conversion failed: {e}")
            raise DataProcessingError(f"XLS → XLSX conversion failed: {e}") from e

    def convert_office_file(self, input_path: str) -> Tuple[str, Any]:
        """
        Detect file extension, convert/load the file,
        and return a tuple of (new_extension, converted_obj).
        Supported extensions: .doc, .docx, .ppt, .pptx, .xls, .xlsx
        """
        self.file_path = self.fileUtils.create_abs_path(input_path, must_exist=True)
        lower = self.file_path.lower()

        if lower.endswith(".doc"):
            self.doc_to_docx()
            return "docx", self.converted_obj

        if lower.endswith(".docx"):
            self.converted_obj = WordDoc(self.file_path)
            return "docx", self.converted_obj

        if lower.endswith(".ppt"):
            self.ppt_to_pptx()
            return "pptx", self.converted_obj

        if lower.endswith(".pptx"):
            self.converted_obj = Presentation(self.file_path)
            return "pptx", self.converted_obj

        if lower.endswith(".xls"):
            self.xls_to_xlsx()
            return "xlsx", self.converted_obj

        if lower.endswith(".xlsx"):
            self.converted_obj = load_workbook(self.file_path)
            return "xlsx", self.converted_obj

        raise ValueError(f"Unsupported extension for '{self.file_path}'")

    def worksheet_to_dataframe(self, ws: Worksheet) -> pd.DataFrame:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return pd.DataFrame()
        return pd.DataFrame(rows[1:], columns=list(rows[0]))  # type: ignore[arg-type]
